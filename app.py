"""
═══════════════════════════════════════════════════════════════════════════
  DASHBOARD REALISASI PENGADAAN PEMERINTAH — INAPROC
  Telkomsel Enterprise | Bid Management — Data Science
  Database: Datamart_Final_Report.db (SQLite)
  ───────────────────────────────────────────────────────────────────────
  streamlit run app.py

  v4.0 — C-Level UI/UX + DuckDB In-Process + Strict K/L Filtering
═══════════════════════════════════════════════════════════════════════════
"""

import streamlit as st
import pandas as pd
import numpy as np
import sqlite3
import duckdb
import seaborn as sns
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import textwrap, os, io, hashlib, re, json, gdown, requests
from datetime import datetime
from difflib import SequenceMatcher
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════════════
# PAGE CONFIG
# ═══════════════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="Realisasi Pengadaan INAPROC — Telkomsel Enterprise",
    page_icon="📊", layout="wide", initial_sidebar_state="expanded")

# ═══════════════════════════════════════════════════════════════════════════
# PASSWORD AUTHENTICATION
# ═══════════════════════════════════════════════════════════════════════════
def _hash(pw):
    return hashlib.sha256(pw.encode()).hexdigest()

def _get_valid_hash():
    try:
        return st.secrets["auth"]["password_hash"]
    except (KeyError, FileNotFoundError):
        return hashlib.sha256("TelkomselEnterprise2025ebpm".encode()).hexdigest()

def check_password():
    if st.session_state.get("authenticated"):
        return True
    st.markdown("""
    <style>
        .login-box{max-width:440px;margin:80px auto;background:#FFF;border-radius:20px;
            padding:48px 40px;text-align:center;box-shadow:0 8px 40px rgba(0,0,0,0.08);
            border-top:5px solid #C8102E;}
        .login-box h2{color:#0D1B2A!important;font-size:22px!important;font-weight:800!important;margin:16px 0 4px!important;}
        .login-box p{color:#5D6D7E!important;font-size:13px!important;margin:0 0 24px!important;}
    </style>
    <div class="login-box">
        <div style="font-size:48px">🔒</div>
        <h2>Telkomsel Enterprise</h2>
        <p>Dashboard Realisasi Pengadaan INAPROC<br>Masukkan password untuk melanjutkan</p>
    </div>""", unsafe_allow_html=True)
    _,col_m,_ = st.columns([1,2,1])
    with col_m:
        password = st.text_input("Password", type="password", key="login_pw", placeholder="Masukkan password...")
        if st.button("🔐 Masuk", use_container_width=True, type="primary"):
            if _hash(password) == _get_valid_hash():
                st.session_state["authenticated"] = True
                st.rerun()
            else:
                st.error("❌ Password salah.")
        st.markdown("<div style='text-align:center;margin-top:20px'><span style='color:#5D6D7E;font-size:11px'>Bid Management — Data Science | 2026</span></div>", unsafe_allow_html=True)
    return False

if not check_password():
    st.stop()

# ═══════════════════════════════════════════════════════════════════════════
# CSS — C-LEVEL PRESENTATION DESIGN SYSTEM v4
# ═══════════════════════════════════════════════════════════════════════════
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600;700;800&family=Inter:wght@400;500;600;700;800&display=swap');

/* ── Base ── */
.stApp{background-color:#F7F9FC;font-family:'Inter',sans-serif;}
.stApp,.stApp p,.stApp span,.stApp div,.stApp label,.stApp li{
    color:#2E4057!important;font-family:'Inter',sans-serif!important;font-size:14px;line-height:1.7;}
.stApp h1,.stApp h2,.stApp h3,.stApp h4,.stApp h5{
    color:#0D1B2A!important;font-family:'DM Sans',sans-serif!important;}
.block-container{padding-top:0;max-width:1480px;padding-left:2rem;padding-right:2rem;}

/* ── Sidebar ── */
section[data-testid="stSidebar"]{
    background:linear-gradient(180deg,#0D1B2A 0%,#1B2A3B 60%,#0D1B2A 100%);
    width:280px;}
section[data-testid="stSidebar"] *{color:#ECF0F1!important;font-family:'Inter',sans-serif!important;}
section[data-testid="stSidebar"] .stRadio label span{color:#ECF0F1!important;font-size:13px!important;font-weight:600!important;}
section[data-testid="stSidebar"] .stRadio div[role="radiogroup"] label[data-selected="true"]{
    background:rgba(200,16,46,0.15)!important;border-left:3px solid #C8102E!important;}
section[data-testid="stSidebar"] .stRadio div[role="radiogroup"] label:hover{background:rgba(255,255,255,0.06)!important;}
section[data-testid="stSidebar"] hr{border-color:rgba(255,255,255,0.08)!important;margin:16px 0!important;}
section[data-testid="stSidebar"] .stButton>button{
    border:1px solid rgba(255,255,255,0.15)!important;
    background:rgba(255,255,255,0.06)!important;
    color:#ECF0F1!important;border-radius:10px!important;}
section[data-testid="stSidebar"] .stButton>button:hover{
    border-color:#C8102E!important;color:#FFFFFF!important;}

/* ── Hero Banner ── */
.hero{
    background:linear-gradient(135deg,#C8102E 0%,#8B0000 45%,#0D1B2A 100%);
    padding:44px 52px;border-radius:24px;margin-bottom:40px;
    box-shadow:0 12px 48px rgba(200,16,46,0.22);position:relative;overflow:hidden;}
.hero::before{content:'';position:absolute;top:-80px;right:-80px;width:360px;height:360px;
    background:rgba(255,255,255,0.03);border-radius:50%;}
.hero::after{content:'';position:absolute;bottom:-60px;left:30%;width:240px;height:240px;
    background:rgba(255,255,255,0.02);border-radius:50%;}
.hero h1{color:#FFF!important;font-family:'DM Sans',sans-serif!important;
    font-size:32px!important;font-weight:800!important;margin:0!important;
    letter-spacing:-0.8px!important;line-height:1.2!important;}
.hero p{color:rgba(255,255,255,0.72)!important;font-size:14px!important;
    margin:10px 0 0!important;font-weight:400!important;letter-spacing:0.2px!important;}

/* ── KPI Cards ── */
.kpi{background:#FFFFFF;border:1.5px solid #E8ECF0;border-top:4px solid #C8102E;
    border-radius:16px;padding:24px 20px;text-align:center;min-height:110px;
    box-shadow:0 4px 20px rgba(13,27,42,0.06);margin-bottom:12px;
    transition:box-shadow 0.2s ease;}
.kpi:hover{box-shadow:0 8px 32px rgba(200,16,46,0.12);}
.kpi .num{color:#0D1B2A!important;font-family:'DM Sans',sans-serif!important;
    font-size:26px;font-weight:800;line-height:1.1;}
.kpi .lab{color:#5D6D7E!important;font-family:'Inter',sans-serif!important;
    font-size:10px;font-weight:600;text-transform:uppercase;
    letter-spacing:1.5px;margin-bottom:6px;}
.kpi .sub{color:#85929E!important;font-family:'Inter',sans-serif!important;
    font-size:11px;margin-top:6px;}

/* ═══ SECTION CARDS ═══ */
.dash-section{
    background:#FFFFFF;border:1.5px solid #E8ECF0;border-radius:20px;
    padding:32px 40px;margin:32px 0;box-shadow:0 2px 16px rgba(13,27,42,0.04);}
.dash-section-red{
    background:#FFFFFF;border:1.5px solid #F5C6CB;border-left:6px solid #C8102E;
    border-radius:0 20px 20px 0;padding:32px 40px;margin:32px 0;
    box-shadow:0 2px 16px rgba(200,16,46,0.05);}
.dash-section-blue{
    background:#FFFFFF;border:1.5px solid #C5D8EA;border-left:6px solid #1B4F72;
    border-radius:0 20px 20px 0;padding:32px 40px;margin:32px 0;
    box-shadow:0 2px 16px rgba(27,79,114,0.05);}
.dash-section-dark{
    background:linear-gradient(135deg,#0D1B2A 0%,#1B2A3B 100%);
    border:1.5px solid #2E4057;border-radius:20px;
    padding:32px 40px;margin:32px 0;box-shadow:0 8px 32px rgba(13,27,42,0.20);}
.dash-section-dark *{color:#FFFFFF!important;}
.dash-section-dark .sub-label{color:#85929E!important;}

/* ── Section Headers ── */
.sec-title{font-family:'DM Sans',sans-serif!important;font-size:22px!important;
    font-weight:700!important;color:#0D1B2A!important;margin:0 0 4px!important;
    letter-spacing:-0.3px!important;}
.sec-subtitle{font-family:'Inter',sans-serif!important;font-size:13px!important;
    color:#5D6D7E!important;margin:0!important;}
.sec-title-white{font-family:'DM Sans',sans-serif!important;font-size:22px!important;
    font-weight:700!important;color:#FFFFFF!important;margin:0 0 4px!important;}
.sec-subtitle-white{font-family:'Inter',sans-serif!important;font-size:13px!important;
    color:#85929E!important;margin:0!important;}

/* ── Chart Container ── */
.chart-card{
    background:#FFFFFF;border:1.5px solid #E8ECF0;border-radius:20px;
    padding:32px 36px;margin:16px 0;box-shadow:0 2px 12px rgba(13,27,42,0.03);}
.chart-title{font-family:'DM Sans',sans-serif!important;font-size:18px!important;
    font-weight:700!important;color:#0D1B2A!important;margin:0 0 4px!important;}
.chart-subtitle{font-family:'Inter',sans-serif!important;font-size:13px!important;
    color:#5D6D7E!important;margin:4px 0 12px!important;}

/* ── Table Container ── */
.table-card{
    background:#FFFFFF;border:1.5px solid #E8ECF0;border-radius:20px;
    padding:28px 32px;margin:16px 0;box-shadow:0 2px 12px rgba(13,27,42,0.03);}
.table-title{font-family:'DM Sans',sans-serif!important;font-size:16px!important;
    font-weight:700!important;color:#0D1B2A!important;margin:0 0 8px!important;}

/* ── Wilayah Region Cards ── */
.rcard{border-radius:16px;padding:24px 28px;margin:12px 0;border:1.5px solid;
    border-left-width:6px;box-shadow:0 2px 12px rgba(0,0,0,0.04);}
.rcard h3{font-family:'DM Sans',sans-serif!important;font-size:20px;font-weight:800;margin:0 0 4px;}
.rcard p{font-family:'Inter',sans-serif!important;font-size:13px;font-weight:500;
    margin:2px 0;color:#2E4057!important;line-height:1.6;}

/* ═══ DOWNLOAD BUTTONS ═══ */
.stDownloadButton>button{
    background:#0D1B2A!important;color:#FFFFFF!important;
    font-family:'Inter',sans-serif!important;font-size:13px!important;
    font-weight:700!important;letter-spacing:0.3px!important;
    border:2px solid #0D1B2A!important;border-radius:12px!important;
    padding:12px 28px!important;box-shadow:0 2px 8px rgba(13,27,42,0.15)!important;
    transition:all 0.25s ease!important;}
.stDownloadButton>button:hover{
    background:#C8102E!important;border-color:#C8102E!important;
    box-shadow:0 6px 20px rgba(200,16,46,0.30)!important;
    transform:translateY(-1px)!important;}

/* ── Tabs ── */
.stTabs [data-baseweb="tab-list"]{background:#F0F4F8;border-radius:10px 10px 0 0;}
.stTabs [data-baseweb="tab"]{font-family:'Inter',sans-serif!important;
    font-weight:700;font-size:13px;padding:12px 28px;color:#5D6D7E!important;}
.stTabs [data-baseweb="tab"][aria-selected="true"]{
    color:#C8102E!important;border-bottom:3px solid #C8102E!important;background:#FFFFFF!important;}
.stTabs [data-baseweb="tab-panel"]{background:#FFFFFF;border:1.5px solid #E8ECF0;
    border-radius:0 16px 16px 16px;padding:28px;}
.streamlit-expanderHeader{font-family:'Inter',sans-serif!important;
    font-size:14px!important;font-weight:700!important;color:#0D1B2A!important;}
details[data-testid="stExpander"]{background:#F7F9FC;border:1.5px solid #E8ECF0;
    border-radius:12px;margin-bottom:10px;}
details[data-testid="stExpander"][open]{border-left:4px solid #C8102E;}
details[data-testid="stExpander"]:hover{background:#F0F4F8;}

/* ── Separator ── */
.sep-line{border:none;border-top:1.5px solid #E8ECF0;margin:36px 0;}
.sep-thick{border:none;border-top:3px solid #C8102E;margin:44px 0;opacity:0.25;}

/* ── Footer ── */
.footer-box{
    background:linear-gradient(135deg,#0D1B2A 0%,#1B2A3B 100%);
    border-radius:20px;padding:32px 40px;text-align:center;
    margin-top:48px;border-top:4px solid #C8102E;
    box-shadow:0 -4px 24px rgba(13,27,42,0.10);}
.stApp .footer-box p,.stApp .footer-box span,.stApp .footer-box div{color:#FFFFFF!important;}
.stApp .footer-box .ft-sub{color:rgba(255,255,255,0.65)!important;}
.stApp .footer-box .ft-dim{color:rgba(255,255,255,0.40)!important;}

/* ── Download Wrapper ── */
.dl-wrap{background:#F0F4F8;border-radius:14px;padding:20px 24px;
    border-left:4px solid #0D1B2A;margin:24px 0;}
.dl-wrap .dl-label{font-size:13px;font-weight:700;color:#0D1B2A!important;margin:0 0 12px;}

#MainMenu{visibility:hidden;} footer{visibility:hidden;}
</style>""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════════
# CONSTANTS
# ═══════════════════════════════════════════════════════════════════════════
DB_NAME = "Create_Lite.db"
GDRIVE_FILE_ID = "1vo4oi_v8ePU6WAPeRmsUG_-bbTsjMmqD"
PARQUET_CACHE = "datamart_cache.parquet"

WILAYAH_MAP = {
    "Aceh":"Sumatera","Sumatera Utara":"Sumatera","Sumatera Barat":"Sumatera",
    "Riau":"Sumatera","Kepulauan Riau":"Sumatera","Jambi":"Sumatera",
    "Sumatera Selatan":"Sumatera","Bangka Belitung":"Sumatera",
    "Bengkulu":"Sumatera","Lampung":"Sumatera",
    "DKI Jakarta":"Jawa","Jawa Barat":"Jawa","Jawa Tengah":"Jawa",
    "DI Yogyakarta":"Jawa","Jawa Timur":"Jawa","Banten":"Jawa",
    "Kalimantan Barat":"Kalimantan","Kalimantan Tengah":"Kalimantan",
    "Kalimantan Selatan":"Kalimantan","Kalimantan Timur":"Kalimantan",
    "Kalimantan Utara":"Kalimantan",
    "Sulawesi Utara":"Sulawesi","Gorontalo":"Sulawesi","Sulawesi Tengah":"Sulawesi",
    "Sulawesi Selatan":"Sulawesi","Sulawesi Barat":"Sulawesi","Sulawesi Tenggara":"Sulawesi",
    "Bali":"Bali NusRa","Nusa Tenggara Barat":"Bali NusRa","Nusa Tenggara Timur":"Bali NusRa",
    "Papua":"Papua Maluku","Papua Barat":"Papua Maluku","Papua Selatan":"Papua Maluku",
    "Papua Tengah":"Papua Maluku","Papua Pegunungan":"Papua Maluku",
    "Papua Barat Daya":"Papua Maluku","Maluku":"Papua Maluku","Maluku Utara":"Papua Maluku",
}

WILAYAH_LIST = ["Sumatera","Jawa","Kalimantan","Sulawesi","Bali NusRa","Papua Maluku"]

W_CFG = {
    "Sumatera":     {"c":"#C8102E","bg":"#FEF0F2","i":"🔴"},
    "Jawa":         {"c":"#1B4F72","bg":"#EBF5FB","i":"🔵"},
    "Kalimantan":   {"c":"#B7770D","bg":"#FEF9E7","i":"🟡"},
    "Sulawesi":     {"c":"#0E6655","bg":"#E8F8F5","i":"🟢"},
    "Bali NusRa":   {"c":"#7B3F00","bg":"#FDF2E9","i":"🟠"},
    "Papua Maluku": {"c":"#4A235A","bg":"#F5EEF8","i":"🟣"},
}

WILAYAH_STRATEGY = {
    "Sumatera":{"tkd":"TKD diarahkan ke revitalisasi sekolah, irigasi, dan koperasi lokal.",
        "dinas":["Dinas Pendidikan","Dinas PUPR","Dinas Pertanian","Dinas Koperasi"],
        "produk":"IoT Smart Farming, Fleet Management, Telkomsel Learning Platform"},
    "Jawa":{"tkd":"Didorong menjadi megalopolis nasional — pusat industri teknologi dan ekonomi kreatif.",
        "dinas":["Diskominfo","Dinas Perindustrian","Dinas Perdagangan","Bappenda"],
        "produk":"Omnichannel, Msight, Tsurvey, IoT Monitoring Management"},
    "Kalimantan":{"tkd":"TKD diarahkan untuk infrastruktur dasar, energi, dan transportasi pendukung IKN.",
        "dinas":["Dinas PUPR","Dinas Perhubungan","Dinas ESDM"],
        "produk":"IoT Smart City, Industrial IoT, IoT Smart Energy Meter"},
    "Sulawesi":{"tkd":"TKD diarahkan untuk sekolah rakyat, irigasi pertanian, dan smart tourism infrastructure.",
        "dinas":["Dinas Pendidikan","Dinas PUPR","Dinas Pariwisata"],
        "produk":"IoT FleetSight, IoT Smart Connectivity, Msight/TSurvey"},
    "Bali NusRa":{"tkd":"TKD diarahkan untuk peningkatan kualitas pendidikan, gizi dan koperasi pariwisata lokal.",
        "dinas":["Dinas Pendidikan","Dinas Kesehatan","Dinas Pariwisata","Dinas Koperasi"],
        "produk":"DigiAds, Msight, Tsurvey, IoT Smart Connectivity"},
    "Papua Maluku":{"tkd":"TKD diarahkan ke pendidikan & kesehatan dasar, pengembangan perikanan & energi terbarukan.",
        "dinas":["Dinas Pendidikan","Dinas Kesehatan","Dinas Perikanan","Dinas ESDM"],
        "produk":"Basic Connectivity, IoT Smart Connectivity, OmniChannel"},
}

DINAS_PATTERNS = {
    "Dinas Pendidikan":  r'(?i)(dinas\s*pendidikan|disdik)',
    "Dinas PUPR":        r'(?i)(dinas\s*(pupr|pekerjaan\s*umum|pu\b|cipta\s*karya))',
    "Dinas Pertanian":   r'(?i)(dinas\s*pertanian|distan|ketahanan\s*pangan)',
    "Dinas Koperasi":    r'(?i)(dinas\s*koperasi|dinkop)',
    "Dinas Kesehatan":   r'(?i)(dinas\s*kesehatan|dinkes)',
    "Dinas Pariwisata":  r'(?i)(dinas\s*pariwisata|dispar)',
    "Dinas Perhubungan": r'(?i)(dinas\s*perhubungan|dishub)',
    "Dinas ESDM":        r'(?i)(dinas\s*(esdm|energi))',
    "Dinas Perindustrian":r'(?i)(dinas\s*perindustrian|disperindag)',
    "Dinas Perdagangan": r'(?i)(dinas\s*perdagangan)',
    "Bappenda":          r'(?i)(bappenda|pendapatan\s*daerah)',
    "Diskominfo":        r'(?i)(diskominfo|dinas\s*komunikasi|kominfo|informatika)',
    "Dinas Perikanan":   r'(?i)(dinas\s*(perikanan|kelautan))',
}
DINAS_COMPILED = {k: re.compile(v) for k, v in DINAS_PATTERNS.items()}

# ── TEMA K/L — 6 Tema Strategis dengan Blacklist Ketat ──
TEMA_KL = {

    "🎓 Penguatan Pendidikan": {
        "kw_inst": [
            r"(?i)(kementerian\s*(pendidikan|riset|teknologi|dikti|dikbud|dikbudristek))",
            r"(?i)(kemen\s*pan|kementerian\s*pendayagunaan\s*aparatur\s*negara|pan\s*rb\b)",
            r"(?i)(lembaga\s*administrasi\s*negara\b|lan\b)",
            r"(?i)(kementerian\s*agama\b|kemenag\b)",
            r"(?i)(bappenas\b|badan\s*perencanaan\s*pembangunan\s*nasional)",
        ],
        "kw_satker": [
            r"(?i)(pendidikan|universitas|politeknik|sekolah|pelatihan|diklat|perguruan\s*tinggi|pesantren|madrasah|aparatur\s*sipil|administrasi\s*negara)",
        ],
        "kw_excl_inst": [
            r"(?i)(kementerian\s*(kesehatan|pertanian|pertahanan|keuangan|sosial|perdagangan|perindustrian|perhubungan|energi|esdm|hukum|komunikasi|komdigi|kominfo|kelautan|transmigrasi|investasi|hilirisasi|pariwisata|koperasi|desa|perumahan|pekerjaan\s*umum))",
            r"(?i)(polri\b|kepolisian)",
            r"(?i)(tni\b|tentara\s*nasional|angkatan\s*(darat|laut|udara)\b)",
            r"(?i)(bnpt\b|bssn\b|bakamla\b)",
            r"(?i)(kementerian\s*imigrasi|kemenkumham\b)",
            r"(?i)(bpjs\b|badan\s*penyelenggara\s*jaminan)",
            r"(?i)(brin\b|badan\s*riset\s*dan\s*inovasi)",
            r"(?i)(bakti\b|bkpm\b)",
        ],
        "color": "#1B4F72", "icon": "🎓",
        "desc": "Kemendikti · KemenPANRB · LAN · Kemenag · BAPPENAS",
    },

    "🏥 Sektor Kesehatan": {
        "kw_inst": [
            r"(?i)(kementerian\s*kesehatan\b|kemenkes\b)",
            r"(?i)(badan\s*riset\s*dan\s*inovasi\s*nasional\b|brin\b)",
            r"(?i)(kementerian\s*sosial\b|kemensos\b)",
            r"(?i)(kemenko\s*(pmk|pembangunan\s*manusia)\b|kementerian\s*koordinator\s*(bidang\s*)?(pmk|pembangunan\s*manusia))",
            r"(?i)(badan\s*gizi\s*nasional\b|bgn\b)",
            r"(?i)(bpjs\b|badan\s*penyelenggara\s*jaminan\s*sosial)",
        ],
        "kw_satker": [
            r"(?i)(kesehatan|rumah\s*sakit\b|rsud\b|puskesmas|farmasi|alat\s*kesehatan|gizi|kesehatan\s*masyarakat|perlindungan\s*sosial|keluarga\s*berencana)",
        ],
        "kw_excl_inst": [
            r"(?i)(kementerian\s*(pendidikan|pertanian|pertahanan|keuangan|perdagangan|perindustrian|perhubungan|energi|esdm|hukum|komunikasi|komdigi|kominfo|kelautan|transmigrasi|investasi|hilirisasi|pariwisata|koperasi|desa|perumahan|pekerjaan\s*umum|agama))",
            r"(?i)(polri\b|kepolisian)",
            r"(?i)(tni\b|tentara\s*nasional|angkatan\s*(darat|laut|udara)\b)",
            r"(?i)(bnpt\b|bssn\b|bakamla\b)",
            r"(?i)(kemenkumham\b|kementerian\s*imigrasi)",
            r"(?i)(bakti\b|bkpm\b)",
            r"(?i)(kemen\s*pan\b|pan\s*rb\b|lembaga\s*administrasi\s*negara\b)",
        ],
        "color": "#C8102E", "icon": "🏥",
        "desc": "Kemenkes · BRIN · Kemensos · KemenkoPMK · BGN · BPJS",
    },

    "🛡️ Pertahanan & Keamanan": {
        "kw_inst": [
            r"(?i)(kementerian\s*pertahanan\b|kemenhan\b)",
            r"(?i)(tentara\s*nasional\s*indonesia\b|tni\b|angkatan\s*(darat|laut|udara)\b|kodam\b|korem\b|kodim\b|koarmada\b|lanud\b|lanal\b)",
            r"(?i)(kepolisian\s*(negara\s*)?republik\s*indonesia\b|polri\b|polda\b|polres\b|polsek\b)",
            r"(?i)(badan\s*nasional\s*penanggulangan\s*terorisme\b|bnpt\b)",
            r"(?i)(badan\s*siber\s*dan\s*sandi\s*negara\b|bssn\b)",
            r"(?i)(badan\s*keamanan\s*laut\b|bakamla\b)",
            r"(?i)(kementerian\s*(hukum|imigrasi|pemasyarakatan)\b|kemenkumham\b|kementerian\s*(hukum\s*dan\s*)?hak\s*asasi\s*manusia)",
        ],
        "kw_satker": [
            r"(?i)(pertahanan|militer|keamanan\s*nasional|intelijen|siber\s*sandi|imigrasi|pemasyarakatan|lapas\b|rutan\b)",
        ],
        "kw_excl_inst": [
            r"(?i)(kementerian\s*(pendidikan|kesehatan|pertanian|keuangan|sosial|perdagangan|perindustrian|perhubungan|energi|esdm|komunikasi|komdigi|kominfo|kelautan|transmigrasi|investasi|hilirisasi|pariwisata|koperasi|desa|perumahan|pekerjaan\s*umum|agama))",
            r"(?i)(bpjs\b|badan\s*penyelenggara\s*jaminan)",
            r"(?i)(brin\b|badan\s*riset\s*dan\s*inovasi)",
            r"(?i)(bakti\b|bkpm\b)",
            r"(?i)(kemen\s*pan\b|pan\s*rb\b|lembaga\s*administrasi\s*negara\b)",
            r"(?i)(bappenas\b)",
            r"(?i)(kejaksaan\s*(agung|republik\s*indonesia)?\b)",
            r"(?i)(mahkamah\s*(agung|konstitusi)\b)",
        ],
        "color": "#2E4057", "icon": "🛡️",
        "desc": "Kemenhan · TNI · Polri · BNPT · BSSN · Bakamla · Kemenkumham",
    },

    "📡 KOMDIGI": {
        "kw_inst": [
            r"(?i)(kementerian\s*komunikasi\s*(dan\s*)?(digital\b|informatika\b))",
            r"(?i)(\bkomdigi\b)",
            r"(?i)(\bkominfo\b)",
            r"(?i)(badan\s*aksesibilitas\s*telekomunikasi\s*dan\s*informasi\b|\bbakti\b)",
        ],
        "kw_satker": [
            r"(?i)(komunikasi\s*(dan\s*)?digital|informatika|telekomunikasi|penyiaran|digital\s*nasional)",
        ],
        "kw_excl_inst": [
            r"(?i)(kementerian\s*(pendidikan|kesehatan|pertanian|pertahanan|keuangan|sosial|perdagangan|perindustrian|perhubungan|energi|esdm|hukum|kelautan|transmigrasi|investasi|hilirisasi|pariwisata|koperasi|desa|perumahan|pekerjaan\s*umum|agama))",
            r"(?i)(polri\b|kepolisian)",
            r"(?i)(tni\b|tentara\s*nasional|angkatan\s*(darat|laut|udara)\b)",
            r"(?i)(bnpt\b|bssn\b|bakamla\b)",
            r"(?i)(kemenkumham\b|kementerian\s*imigrasi)",
            r"(?i)(bpjs\b|badan\s*penyelenggara\s*jaminan)",
            r"(?i)(brin\b|badan\s*riset\s*dan\s*inovasi)",
            r"(?i)(bkpm\b|badan\s*koordinasi\s*penanaman)",
            r"(?i)(kemen\s*pan\b|pan\s*rb\b|lembaga\s*administrasi\s*negara\b)",
            r"(?i)(bappenas\b)",
            r"(?i)(diskominfo\b|dinas\s*komunikasi)",
        ],
        "color": "#4A235A", "icon": "📡",
        "desc": "HANYA: Kementerian Komunikasi dan Digital (KOMDIGI) & BAKTI",
    },

    "⚡ Subsidi Energi & Hilirisasi": {
        "kw_inst": [
            r"(?i)(kementerian\s*energi\s*(dan\s*)?sumber\s*daya\s*mineral\b|kemen\s*esdm\b)",
        ],
        "kw_satker": [
            r"(?i)(energi|hilirisasi|esdm\b|minyak\s*(dan\s*)?gas|gas\s*bumi|pertambangan|mineral\b|nikel|bauksit|batubara|kilang|pembangkit\s*listrik)",
        ],
        "kw_excl_inst": [
            r"(?i)(kementerian\s*(pendidikan|kesehatan|pertanian|pertahanan|keuangan|sosial|perdagangan|perindustrian|perhubungan|komunikasi|komdigi|kominfo|hukum|kelautan|transmigrasi|investasi|hilirisasi|pariwisata|koperasi|desa|perumahan|pekerjaan\s*umum|agama))",
            r"(?i)(polri\b|kepolisian)",
            r"(?i)(tni\b|tentara\s*nasional|angkatan\s*(darat|laut|udara)\b)",
            r"(?i)(bnpt\b|bssn\b|bakamla\b)",
            r"(?i)(kemenkumham\b|kementerian\s*imigrasi)",
            r"(?i)(bpjs\b|badan\s*penyelenggara\s*jaminan)",
            r"(?i)(brin\b|badan\s*riset\s*dan\s*inovasi)",
            r"(?i)(bakti\b|bkpm\b|badan\s*koordinasi\s*penanaman)",
            r"(?i)(kemen\s*pan\b|pan\s*rb\b|lembaga\s*administrasi\s*negara\b)",
            r"(?i)(bappenas\b)",
        ],
        "color": "#B7770D", "icon": "⚡",
        "desc": "HANYA: Kementerian ESDM",
    },

    "🏗️ Pembangunan Ekonomi & Infrastruktur": {
        "kw_inst": [
            r"(?i)(kementerian\s*pekerjaan\s*umum\b|kemen\s*pu\b|direktorat\s*jenderal\s*(bina\s*marga|cipta\s*karya|sumber\s*daya\s*air|pembiayaan\s*infrastruktur))",
            r"(?i)(kementerian\s*perumahan\s*(dan\s*)?kawasan\s*permukiman\b)",
            r"(?i)(kementerian\s*perhubungan\b|kemenhub\b)",
            r"(?i)(kementerian\s*keuangan\b|kemenkeu\b|bendahara\s*(umum\s*)?negara\b|direktorat\s*jenderal\s*(pajak|bea\s*cukai|perbendaharaan|anggaran|kekayaan\s*negara))",
            r"(?i)(kementerian\s*hilirisasi\b|kementerian\s*investasi\b|\bbkpm\b|badan\s*koordinasi\s*penanaman\s*modal\b)",
        ],
        "kw_satker": [
            r"(?i)(perhubungan|transportasi|bandara|pelabuhan|perkeretaapian|lalu\s*lintas\s*jalan|pajak|bea\s*cukai|perbendaharaan|kekayaan\s*negara|pekerjaan\s*umum|bina\s*marga|cipta\s*karya|sumber\s*daya\s*air|perumahan|kawasan\s*permukiman|investasi|modal)",
        ],
        "kw_excl_inst": [
            r"(?i)(kementerian\s*(pendidikan|kesehatan|pertanian|pertahanan|sosial|perdagangan|perindustrian|energi|esdm|komunikasi|komdigi|kominfo|hukum|kelautan|transmigrasi|pariwisata|koperasi|desa|agama))",
            r"(?i)(polri\b|kepolisian)",
            r"(?i)(tni\b|tentara\s*nasional|angkatan\s*(darat|laut|udara)\b)",
            r"(?i)(bnpt\b|bssn\b|bakamla\b)",
            r"(?i)(kemenkumham\b|kementerian\s*imigrasi)",
            r"(?i)(bpjs\b|badan\s*penyelenggara\s*jaminan)",
            r"(?i)(brin\b|badan\s*riset\s*dan\s*inovasi)",
            r"(?i)(bakti\b)",
            r"(?i)(kemen\s*pan\b|pan\s*rb\b|lembaga\s*administrasi\s*negara\b)",
            r"(?i)(bappenas\b)",
        ],
        "color": "#0E6655", "icon": "🏗️",
        "desc": "Kemen PU · Kemen Perumahan · Kemenhub · Kemenkeu · Kemen Hilirisasi & Investasi/BKPM",
    },
}

# Compile all regex patterns (including blacklist)
for _t in TEMA_KL.values():
    _t["_re_inst"]      = [re.compile(p) for p in _t["kw_inst"]]
    _t["_re_satker"]    = [re.compile(p) for p in _t["kw_satker"]]
    _t["_re_excl_inst"] = [re.compile(p) for p in _t.get("kw_excl_inst", [])]

# ═══════════════════════════════════════════════════════════════════════════
# ICT MEGA-REGEX
# ═══════════════════════════════════════════════════════════════════════════
ICT_WHITELIST = [
    r'\binternet\b',r'\bbandwidth\b',r'\bfiber\s*optik?\b',r'\bjaringan\b',
    r'\bwifi\b',r'\bwi-fi\b',r'\bhotspot\b',r'\bmpls\b',r'\bvpn\b',r'\bsd-wan\b',
    r'\bbroadband\b',r'\btelekomunikasi\b',r'\bfttx?\b',r'\bdata\s*center\b',
    r'\bserver\b(?!.*makanan)',r'\bkomputer\b',r'\blaptop\b',r'\bnotebook\b',
    r'\bprinter\b',r'\bscanner\b',r'\bups\b',
    r'\bswitch\b(?!.*listrik)',r'\brouter\b',r'\bfirewall\b',r'\baccess\s*point\b',
    r'\bstorage\b',r'\brack\b(?!.*sepeda)',r'\bcctv\b',r'\bip\s*camera\b',
    r'\bnetwork\b',r'\binfrastruktur\s*(it|ti|ict|teknologi)\b',
    r'\baplikasi\b',r'\bsoftware\b',r'\bperangkat\s*lunak\b',r'\blisens[i]\b',
    r'\bsistem\s*informasi\b',r'\be-gov\w*\b',r'\bwebsite\b',r'\bportal\b',
    r'\bcloud\b',r'\bsaas\b',r'\berp\b',r'\bdatabase\b',r'\bbig\s*data\b',
    r'\bmachine\s*learning\b',r'\bcyber\s*security\b',r'\bkeamanan\s*siber\b',
    r'\biot\b',r'\bsmart\s*(city|village|building|farming|meter)\b',
    r'\bsensor\b(?!.*gas\s*lpg)',r'\btelemetri\b',r'\bsurveillance\b',
    r'\bsim\s*card\b',r'\bpulsa\b',r'\bpaket\s*data\b',r'\bsms\s*(gateway|blast)\b',
    r'\bvoip\b',r'\bip\s*phone\b',r'\bpabx\b',r'\bvideo\s*conference\b',
    r'\bdigital\s*(signage|marketing|transform)\b',r'\bomnichannel\b',
    r'\bpc\b(?!.*pcs)',r'\blan\b(?!.*lain)',r'\bwan\b(?!.*wan)',
]
ICT_BLACKLIST = [
    r'\bgaji\b',r'\bhonor\w*\b',r'\btunjangan\b',r'\bmakanan\b',r'\bminuman\b',
    r'\bjas\s*hujan\b',r'\bseragam\b',r'\bbaju\b',r'\bsepatu\b',
    r'\bkonstruksi\b(?!.*(smart|iot|sensor))',r'\bjalan\b(?!.*(smart|monitoring))',
    r'\bjembatan\b',r'\birigasi\b(?!.*(smart|iot))',
    r'\bpengolah\w*\s*sampah\b',r'\bpetugas\b',r'\bcaraka\b',r'\bnormalisasi\b',
    r'\brestorasi\b',r'\bperawat\s*taman\b',
    r'\bpemeliharaan\b(?!.*(server|jaringan|it|network))',
]
_ICT_WL_SIMPLE, _ICT_WL_LOOKAHEAD = [], []
for p in ICT_WHITELIST:
    (_ICT_WL_LOOKAHEAD if '(?!' in p else _ICT_WL_SIMPLE).append(
        re.compile(p, re.IGNORECASE) if '(?!' in p else p)
_ICT_BL_SIMPLE, _ICT_BL_LOOKAHEAD = [], []
for p in ICT_BLACKLIST:
    (_ICT_BL_LOOKAHEAD if '(?!' in p else _ICT_BL_SIMPLE).append(
        re.compile(p, re.IGNORECASE) if '(?!' in p else p)
_ICT_WL_MEGA = re.compile('|'.join(_ICT_WL_SIMPLE), re.IGNORECASE) if _ICT_WL_SIMPLE else None
_ICT_BL_MEGA = re.compile('|'.join(_ICT_BL_SIMPLE), re.IGNORECASE) if _ICT_BL_SIMPLE else None

# ═══════════════════════════════════════════════════════════════════════════
# FORMAT HELPERS
# ═══════════════════════════════════════════════════════════════════════════
def fmt_rp(v):
    if pd.isna(v) or v == 0: return "Rp 0"
    a = abs(v)
    if a >= 1e12: return f"Rp {v/1e12:,.2f} T"
    if a >= 1e9:  return f"Rp {v/1e9:,.2f} M"
    if a >= 1e6:  return f"Rp {v/1e6:,.1f} Jt"
    return f"Rp {v:,.0f}"

def fmt_s(v):
    if pd.isna(v) or v == 0: return "0"
    a = abs(v)
    if a >= 1e12: return f"{v/1e12:.1f}T"
    if a >= 1e9:  return f"{v/1e9:.1f}M"
    if a >= 1e6:  return f"{v/1e6:.0f}Jt"
    return f"{v:,.0f}"

def fmt_n(v):
    if pd.isna(v): return "0"
    return f"{int(v):,}".replace(",", ".")

def kpi(lb, vl, sb=""):
    s = f'<div class="sub">{sb}</div>' if sb else ""
    return f'<div class="kpi"><div class="lab">{lb}</div><div class="num">{vl}</div>{s}</div>'

# ═══════ UI WRAPPERS ═══════

def section_open(title, subtitle="", style="red"):
    """Open a visual section card with title."""
    cls = {"red":"dash-section-red","blue":"dash-section-blue","dark":"dash-section-dark","plain":"dash-section"}[style]
    tcls = "sec-title-white" if style == "dark" else "sec-title"
    scls = "sec-subtitle-white" if style == "dark" else "sec-subtitle"
    sub = f'<p class="{scls}">{subtitle}</p>' if subtitle else ""
    st.markdown(f'<div class="{cls}"><p class="{tcls}">{title}</p>{sub}</div>', unsafe_allow_html=True)

def chart_card_open(title, subtitle=""):
    sub = f'<p class="chart-subtitle">{subtitle}</p>' if subtitle else ""
    st.markdown(f'<div class="chart-card"><p class="chart-title">{title}</p>{sub}', unsafe_allow_html=True)

def chart_card_close():
    st.markdown('</div>', unsafe_allow_html=True)

def table_card_open(title):
    st.markdown(f'<div class="table-card"><p class="table-title">{title}</p>', unsafe_allow_html=True)

def table_card_close():
    st.markdown('</div>', unsafe_allow_html=True)

def dl_wrap_open(label="📥 Unduh Data"):
    st.markdown(f'<div class="dl-wrap"><p class="dl-label">{label}</p>', unsafe_allow_html=True)

def dl_wrap_close():
    st.markdown('</div>', unsafe_allow_html=True)

def separator():
    st.markdown('<hr class="sep-line">', unsafe_allow_html=True)

def separator_thick():
    st.markdown('<hr class="sep-thick">', unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT
# ═══════════════════════════════════════════════════════════════════════════
def _safe_sheet(name):
    return re.sub(r'[\*\?/\\\[\]:]', '', str(name))[:31].strip() or "Data"

def to_excel_styled(df, sheet_name="Data"):
    sheet_name = _safe_sheet(sheet_name)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name=sheet_name)
        ws = w.sheets[sheet_name]
        hf = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        hfl = PatternFill(start_color="C41920", end_color="C41920", fill_type="solid")
        ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="DDDDDD")
        brd = Border(top=thin, left=thin, right=thin, bottom=thin)
        bf = Font(name="Calibri", size=10)
        af = PatternFill(start_color="FFF5F5", end_color="FFF5F5", fill_type="solid")
        for ci in range(1, len(df.columns)+1):
            c = ws.cell(row=1, column=ci); c.font=hf; c.fill=hfl; c.alignment=ha; c.border=brd
            ws.column_dimensions[get_column_letter(ci)].width = max(15, min(45, len(str(c.value or ""))+4))
        for ri in range(2, len(df)+2):
            for ci in range(1, len(df.columns)+1):
                c = ws.cell(row=ri, column=ci); c.font=bf; c.border=brd
                c.alignment = Alignment(vertical="center", wrap_text=True)
                if ri % 2 == 0: c.fill = af
        ws.auto_filter.ref = ws.dimensions; ws.freeze_panes = "A2"
    return buf.getvalue()

# ═══════════════════════════════════════════════════════════════════════════
# FUZZY UNCENSORING
# ═══════════════════════════════════════════════════════════════════════════
def _build_clean_names_index(all_names):
    clean = [n for n in all_names if '*' not in n and len(n) > 2]
    index = {}
    for n in clean:
        key = (n[0].upper(), len(n))
        index.setdefault(key, []).append(n)
    return clean, index

def _uncensor_name(censored, clean_index):
    try:
        pat_str = ""
        for ch in censored:
            if ch == '*': pat_str += '.'
            elif ch in r'\.[](){}+?^$|': pat_str += '\\' + ch
            else: pat_str += ch
        pat_re = re.compile(f'^{pat_str}$', re.IGNORECASE)
        first_char = censored[0].upper() if censored[0] != '*' else None
        candidates = []
        if first_char:
            bucket = clean_index.get((first_char, len(censored)), [])
            candidates = [n for n in bucket if pat_re.match(n)]
        if not candidates and first_char:
            for delta in [-1, 1]:
                bucket = clean_index.get((first_char, len(censored) + delta), [])
                candidates.extend(n for n in bucket if pat_re.match(n))
        if candidates: return candidates[0]
    except re.error: pass
    first_char = censored[0].upper() if censored[0] != '*' else None
    best_score, best_match = 0.0, None
    if first_char:
        for delta in range(-2, 3):
            bucket = clean_index.get((first_char, len(censored) + delta), [])
            for cn in bucket:
                score = SequenceMatcher(None, censored.upper(), cn.upper()).ratio()
                if score > best_score: best_score = score; best_match = cn
    return best_match if best_score >= 0.6 else censored

def _build_uncensor_mapping(all_names):
    censored_names = [n for n in all_names if '*' in n]
    if not censored_names: return {}
    _, clean_index = _build_clean_names_index(all_names)
    mapping = {}
    for cn in censored_names:
        result = _uncensor_name(cn, clean_index)
        if result != cn: mapping[cn] = result
    return mapping

def uncensor_for_chart(agg_df, uncensor_map, name_col="Nama_Display"):
    if len(agg_df) == 0 or not uncensor_map: return agg_df
    mask = agg_df[name_col].str.contains(r'\*', na=False)
    if not mask.any(): return agg_df
    result = agg_df.copy()
    for idx in result[mask].index:
        old = result.at[idx, name_col]
        result.at[idx, name_col] = uncensor_map.get(old, old)
    return result

# ═══════════════════════════════════════════════════════════════════════════
# DATA LOADING
# ═══════════════════════════════════════════════════════════════════════════

# ═══════════════════════════════════════════════════════════════════════════
# ★ TURSO CONNECTION — libsql (primary) → HTTP chunked (fallback)
# ═══════════════════════════════════════════════════════════════════════════
TURSO_URL = "libsql://datamart-jidiyosua.aws-eu-west-1.turso.io"
TURSO_TOKEN = "eyJhbGciOiJFZERTQSIsInR5cCI6IkpXVCJ9.eyJqdGkiOiJ1VVFJSUN6aEVmR1NKNWJaSVBGSkZ3In0.a1ZPmbTHxRsFfh0MoNHdR6_jL20YZI72uWP_n02oOkTNVsuPeaj6VO8br57QD2IusVvT6CL2QBXKk9s-KsU3AQ"
try:
    TURSO_URL = st.secrets["turso"]["url"]
    TURSO_TOKEN = st.secrets["turso"]["token"]
except:
    pass


class _TursoResult:
    def __init__(self, rows):
        self._rows = rows
    def fetchall(self):
        return self._rows
    def fetchone(self):
        return self._rows[0] if self._rows else None


class TursoHTTP:
    """Turso via HTTP API — chunked fetch for large queries."""
    CHUNK = 100_000

    def __init__(self, url, token):
        self._base = url.replace("libsql://", "https://").rstrip("/")
        self._token = token

    def _post(self, sql, timeout=300):
        r = requests.post(
            f"{self._base}/v2/pipeline",
            headers={"Authorization": f"Bearer {self._token}", "Content-Type": "application/json"},
            json={"requests": [{"type": "execute", "stmt": {"sql": sql}}, {"type": "close"}]},
            timeout=timeout
        )
        if r.status_code != 200:
            raise Exception(f"Turso HTTP {r.status_code}: {r.text[:300]}")
        data = r.json()
        result = data.get("results", [{}])[0]
        if result.get("type") == "error":
            raise Exception(f"Turso SQL: {result.get('error', {}).get('message', '')}")
        return result.get("response", {}).get("result", {})

    @staticmethod
    def _parse_rows(raw_rows):
        parsed = []
        for row in raw_rows:
            vals = []
            for cell in row:
                if cell is None: vals.append(None)
                elif isinstance(cell, dict): vals.append(cell.get("value"))
                else: vals.append(cell)
            parsed.append(tuple(vals))
        return parsed

    def execute(self, sql):
        """Small queries (PRAGMA, COUNT, SELECT 1, etc)."""
        resp = self._post(sql, timeout=60)
        return _TursoResult(self._parse_rows(resp.get("rows", [])))

    def execute_chunked(self, sql, total_hint=0):
        """Large SELECT — fetch in 100K-row chunks to avoid timeout."""
        all_rows = []
        offset = 0
        while True:
            chunk_sql = f"{sql} LIMIT {self.CHUNK} OFFSET {offset}"
            resp = self._post(chunk_sql, timeout=300)
            rows = self._parse_rows(resp.get("rows", []))
            all_rows.extend(rows)
            if total_hint > 0:
                pct = min(100, len(all_rows) / total_hint * 100)
                st.toast(f"Loading: {len(all_rows):,}/{total_hint:,} ({pct:.0f}%)")
            if len(rows) < self.CHUNK:
                break
            offset += self.CHUNK
        return _TursoResult(all_rows)


def _connect_db():
    """Try: 1) libsql (Linux/Streamlit Cloud) 2) HTTP API 3) local SQLite."""
    if TURSO_URL and TURSO_TOKEN:
        try:
            import libsql_experimental as libsql
            conn = libsql.connect(TURSO_URL, auth_token=TURSO_TOKEN)
            conn.execute("SELECT 1")
            return conn, "Turso (libsql)"
        except ImportError:
            pass
        except Exception:
            pass
        try:
            conn = TursoHTTP(TURSO_URL, TURSO_TOKEN)
            conn.execute("SELECT 1")
            return conn, "Turso (HTTP)"
        except Exception as e:
            st.warning(f"Turso unavailable: {e}")
    import sqlite3 as sq3
    sd = os.path.dirname(os.path.abspath(__file__))
    db_path = os.path.join(sd, DB_NAME)
    if os.path.exists(db_path):
        return sq3.connect(db_path), f"Local: {DB_NAME}"
    return None, ""

def _find_db():
    sd = os.path.dirname(os.path.abspath(__file__))
    db_path = os.path.join(sd, DB_NAME)
    if os.path.exists(db_path): return db_path
    try:
        url = f"https://drive.google.com/uc?id={GDRIVE_FILE_ID}"
        gdown.download(url, db_path, quiet=False)
        if os.path.exists(db_path): return db_path
    except Exception as e:
        st.error(f"❌ Gagal download database dari Google Drive: {e}")
    return None

VENDOR_MAPPING_FILE = "vendor_mapping.json"
def _load_vendor_mapping():
    sd = os.path.dirname(os.path.abspath(__file__))
    for base in [sd, ".", os.getcwd()]:
        p = os.path.join(base, VENDOR_MAPPING_FILE)
        if os.path.exists(p):
            with open(p, "r", encoding="utf-8") as f: return json.load(f)
    return {}

def _vectorized_ict_fast(nama_paket_series):
    text = nama_paket_series.fillna("").str.lower()
    result = pd.Series("Non-ICT", index=text.index, dtype="object")
    bl = pd.Series(False, index=text.index)
    if _ICT_BL_MEGA: bl |= text.str.contains(_ICT_BL_MEGA, na=False)
    for pat in _ICT_BL_LOOKAHEAD: bl |= text.str.contains(pat, na=False)
    wl = pd.Series(False, index=text.index)
    if _ICT_WL_MEGA: wl |= text.str.contains(_ICT_WL_MEGA, na=False)
    for pat in _ICT_WL_LOOKAHEAD: wl |= text.str.contains(pat, na=False)
    result[wl & ~bl] = "ICT"
    return result

def _get_parquet_path():
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), PARQUET_CACHE)

@st.cache_data(show_spinner="🔄 Memuat database…")
def load_and_process():
    parquet_path = _get_parquet_path()
    if os.path.exists(parquet_path):
        try:
            df = pd.read_parquet(parquet_path)
            if all(c in df.columns for c in ["Nama_Pemenang","Pagu_Rp","Sektor","Wilayah","Provinsi","is_pemda","Nama_Display"]):
                return df, len(df), 0, None, "Parquet Cache"
        except Exception: pass

    # ★ TURSO (primary) → Local SQLite (fallback)
    conn, db_source = _connect_db()
    if conn is None:
        return None, 0, 0, f"Database tidak tersedia. Set Turso secrets atau letakkan {DB_NAME} lokal.", ""

    try:
        # Detect table
        tables = [r[0] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()]
        if not tables:
            return None, 0, 0, "Database kosong.", db_source
        tbl = tables[0]

        # Detect columns
        pragma = conn.execute(f"PRAGMA table_info([{tbl}])").fetchall()
        db_columns = [row[1] for row in pragma]

        ESSENTIAL = ["Nama_Paket","Pagu_Rp","Instansi_Pembeli","Satuan_Kerja","Lokasi","Nama_Pemenang"]
        OPTIONAL  = ["ID_RUP","Kategori_Paket","Metode_Pemilihan","Jenis_Pengadaan","Total_Pelaksanaan_Rp","Sumber_Data","Prediksi_Nama"]
        missing = [c for c in ESSENTIAL if c not in db_columns]
        if missing:
            return None, 0, 0, f"Kolom essential tidak ditemukan: {missing}", db_source

        select_cols = list(dict.fromkeys(c for c in ESSENTIAL + OPTIONAL if c in db_columns))
        cols_sql = ", ".join([f'[{c}]' for c in select_cols])

        # ★ COUNT + SELECT — works for both libsql and sqlite3
        total_rows = conn.execute(f"SELECT COUNT(*) FROM [{tbl}]").fetchone()[0]

        sql = (f"SELECT {cols_sql} FROM [{tbl}] WHERE [Nama_Pemenang] IS NOT NULL AND TRIM([Nama_Pemenang]) != '' "
               f"AND CAST([Pagu_Rp] AS REAL) > 0")

        # ★ Use chunked fetch for HTTP API (large data), normal for libsql/sqlite3
        if isinstance(conn, TursoHTTP):
            rows = conn.execute_chunked(sql, total_hint=total_rows).fetchall()
        else:
            rows = conn.execute(sql).fetchall()
        df = pd.DataFrame(rows, columns=select_cols)
    except Exception as e:
        return None, 0, 0, f"Error query database ({db_source}): {e}", db_source
    for c in ESSENTIAL + OPTIONAL:
        if c not in df.columns: df[c] = ""
    df["Pagu_Rp"] = pd.to_numeric(df["Pagu_Rp"], errors="coerce").fillna(0)
    df["Total_Pelaksanaan_Rp"] = pd.to_numeric(df.get("Total_Pelaksanaan_Rp", 0), errors="coerce").fillna(0)
    df = df[df["Pagu_Rp"] > 0].reset_index(drop=True)
    n_matched = 0
    mapping = _load_vendor_mapping()
    if mapping:
        mask = df["Nama_Pemenang"].isin(mapping.keys()); n_matched = mask.sum()
        if n_matched > 0: df.loc[mask, "Nama_Pemenang"] = df.loc[mask, "Nama_Pemenang"].map(mapping)
    df["Sektor"] = _vectorized_ict_fast(df["Nama_Paket"])
    df["Provinsi"] = df["Lokasi"].str.split(",").str[0].str.strip().fillna("Lainnya")
    df["Wilayah"] = df["Provinsi"].map(WILAYAH_MAP).fillna("Lainnya")
    inst = df["Instansi_Pembeli"].fillna("").str.strip()
    df["is_pemda"] = inst.str.startswith("Kab.") | inst.str.startswith("Kota ") | inst.str.startswith("Provinsi ")
    for col in ["Sektor","Wilayah","Provinsi","Metode_Pemilihan","Jenis_Pengadaan"]:
        if col in df.columns: df[col] = df[col].astype("category")

    # ★ Prediksi_Nama → Nama_Display (untuk grafik)
    if "Prediksi_Nama" not in df.columns:
        df["Prediksi_Nama"] = ""
    _pred = df["Prediksi_Nama"].fillna("").str.strip()
    _has_star = df["Nama_Pemenang"].str.contains(r"\*", na=False)
    _valid_pred = (_pred != "") & (_pred != "nan")
    df["Nama_Display"] = df["Nama_Pemenang"]
    df.loc[_has_star & _valid_pred, "Nama_Display"] = df.loc[_has_star & _valid_pred, "Prediksi_Nama"]

    try: df.to_parquet(parquet_path, index=False, compression="snappy")
    except Exception: pass
    return df, total_rows, n_matched, None, db_source

# ═══════════════════════════════════════════════════════════════════════════
# ★ DUCKDB IN-PROCESS — Cached connection for fast aggregation
# ═══════════════════════════════════════════════════════════════════════════
@st.cache_resource
def get_duckdb_conn():
    """DuckDB in-process connection — dibuat sekali, dipakai semua sesi."""
    return duckdb.connect(database=":memory:")

def register_df(con, df, table_name="main_df"):
    """Register/refresh DataFrame ke DuckDB. Dipanggil setelah df berubah."""
    con.execute(f"DROP VIEW IF EXISTS {table_name}")
    con.register(table_name, df)

# ═══════════════════════════════════════════════════════════════════════════
# CHART FUNCTIONS — Redesigned with C-Level palette
# ═══════════════════════════════════════════════════════════════════════════
_CHART_RC = {
    "font.family":        "DejaVu Sans",
    "axes.facecolor":     "#FFFFFF",
    "figure.facecolor":   "#FFFFFF",
    "axes.spines.top":    False,
    "axes.spines.right":  False,
    "axes.spines.left":   False,
    "axes.spines.bottom": True,
    "axes.edgecolor":     "#E8ECF0",
    "axes.linewidth":     1.2,
    "grid.color":         "#F0F4F8",
    "grid.linewidth":     0.8,
    "xtick.labelsize":    11,
    "ytick.labelsize":    12,
    "xtick.color":        "#5D6D7E",
    "ytick.color":        "#0D1B2A",
}

def _blend_color(hex_color, alpha, bg=(247, 249, 252)):
    """Blend hex_color with bg at given alpha (0-1)."""
    base = hex_color.lstrip('#')
    r0, g0, b0 = int(base[:2], 16), int(base[2:4], 16), int(base[4:6], 16)
    r = int(r0 * alpha + bg[0] * (1 - alpha))
    g = int(g0 * alpha + bg[1] * (1 - alpha))
    b = int(b0 * alpha + bg[2] * (1 - alpha))
    return f"#{min(255,r):02x}{min(255,g):02x}{min(255,b):02x}"

def chart_top20(df_agg, title, subtitle, accent_color, semesta=None, figsize=(18, 10),
                val_col="Total_Pagu", name_col="Nama_Display"):
    sns.set_theme(style="white", rc=_CHART_RC)
    d = df_agg.head(20).copy().reset_index(drop=True)
    n = len(d)
    if n == 0:
        fig, ax = plt.subplots(figsize=(8, 3))
        ax.text(0.5, 0.5, "Tidak ada data", ha="center", va="center", fontsize=16, color="#5D6D7E")
        ax.axis("off"); return fig
    fig_h = max(10, n * 0.62)
    fig, ax = plt.subplots(figsize=(figsize[0], fig_h))
    fig.patch.set_facecolor("#FFFFFF")
    ax.set_facecolor("#FFFFFF")
    # Gradient palette: full opacity rank 1 → 40% opacity rank 20
    palette = []
    for i in range(n):
        alpha = 1.0 - (i / max(n - 1, 1)) * 0.60
        palette.append(_blend_color(accent_color, alpha))
    y_pos = list(range(n-1, -1, -1))
    ax.barh(y_pos, d[val_col], color=palette, height=0.58, edgecolor="white", linewidth=1.5, zorder=3)
    mx = d[val_col].max()
    for i, (_, row) in enumerate(d.iterrows()):
        y = n - 1 - i; val = row[val_col]
        pct = f"  ({val/semesta*100:.1f}%)" if semesta and semesta > 0 else ""
        line1 = f"{fmt_rp(val)}{pct}"
        parts = []
        if "Jumlah_Paket" in row.index and row["Jumlah_Paket"] > 0: parts.append(f"{int(row['Jumlah_Paket'])} paket")
        if "Instansi_Unik" in row.index and row["Instansi_Unik"] > 0: parts.append(f"{int(row['Instansi_Unik'])} instansi")
        if "Satker_Unik" in row.index and row["Satker_Unik"] > 0: parts.append(f"{int(row['Satker_Unik'])} satker")
        line2 = "  •  ".join(parts)
        ax.text(val + mx*0.010, y + 0.14, line1, va="center", ha="left", fontsize=12, fontweight="bold", color="#0D1B2A")
        if line2: ax.text(val + mx*0.010, y - 0.18, line2, va="center", ha="left", fontsize=10.5, fontweight="600", color="#5D6D7E")
    ax.set_yticks(y_pos)
    ax.set_yticklabels(["\n".join(textwrap.wrap(str(nm), 32)) for nm in d[name_col]], fontsize=12, fontweight="bold", color="#0D1B2A")
    ax.set_title(title, fontsize=20, fontweight="800", color="#0D1B2A", loc="left", pad=28)
    if subtitle: ax.text(0, 1.04, subtitle, transform=ax.transAxes, fontsize=13, color="#5D6D7E", ha="left")
    if semesta:
        ax.text(1.0, -0.06, f"SEMESTA: {fmt_rp(semesta)}",
                transform=ax.transAxes, fontsize=12, fontweight="bold",
                color="#C8102E", ha="right", va="top",
                bbox=dict(boxstyle="round,pad=0.5", facecolor="#FEF0F2", edgecolor="#C8102E", linewidth=1.5, alpha=0.9))
    ax.set_xlabel(""); ax.set_ylabel("")
    ax.tick_params(axis="x", labelsize=11, labelcolor="#5D6D7E")
    ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: fmt_s(x)))
    ax.set_xlim(0, mx * 1.62)
    ax.grid(axis="x", alpha=0.5, linestyle="--", color="#F0F4F8", zorder=0)
    for sp in ["top","right","left"]: ax.spines[sp].set_visible(False)
    ax.spines["bottom"].set_color("#E8ECF0")
    ax.spines["bottom"].set_linewidth(1.2)
    plt.tight_layout(pad=2.5)
    return fig

def chart_donut(val_ict, val_non, cnt_ict, cnt_non, label=""):
    total = val_ict + val_non
    fig, ax = plt.subplots(figsize=(6, 6.2))
    fig.patch.set_facecolor("#FFFFFF"); ax.set_facecolor("#FFFFFF")
    if total == 0:
        ax.text(0.5, 0.5, "Tidak ada data", ha="center", va="center", fontsize=13, color="#5D6D7E", transform=ax.transAxes)
        ax.axis("off"); return fig
    pct_ict = val_ict/total*100
    sizes, colors, labels = [], [], []
    if val_ict > 0:
        sizes.append(val_ict); colors.append("#1B4F72")
        labels.append(f"ICT — {fmt_rp(val_ict)} ({pct_ict:.1f}%) • {fmt_n(cnt_ict)} pemenang")
    if val_non > 0:
        sizes.append(val_non); colors.append("#0E6655")
        labels.append(f"Non-ICT — {fmt_rp(val_non)} ({100-pct_ict:.1f}%) • {fmt_n(cnt_non)} pemenang")
    wedges, _, autotexts = ax.pie(
        sizes, colors=colors, autopct=lambda p: f"{p:.1f}%", startangle=90, pctdistance=0.75,
        wedgeprops=dict(width=0.40, edgecolor="white", linewidth=4),
        textprops=dict(color="white", fontsize=14, fontweight="bold"))
    for at in autotexts: at.set_fontsize(14); at.set_fontweight("800")
    ax.text(0, 0.10, fmt_rp(total), ha="center", va="center", fontsize=16, fontweight="800", color="#0D1B2A")
    ax.text(0, -0.08, label, ha="center", va="center", fontsize=11, fontweight="600", color="#5D6D7E")
    lg = ax.legend(wedges, labels, loc="lower center", bbox_to_anchor=(0.5, -0.14), fontsize=10, frameon=False, ncol=1)
    for t in lg.get_texts(): t.set_color("#0D1B2A")
    ax.set_aspect("equal"); ax.axis("off")
    plt.subplots_adjust(bottom=0.18, top=0.96)
    return fig

def chart_heatmap(df_agg, si_col, seg_col, val_col, title, cmap="Blues"):
    top = df_agg.groupby(si_col)[val_col].sum().sort_values(ascending=False).head(15).index
    h = df_agg[df_agg[si_col].isin(top)].groupby([si_col, seg_col])[val_col].sum().unstack(fill_value=0)
    h["_t"] = h.sum(axis=1); h = h.sort_values("_t", ascending=False).drop("_t", axis=1)
    if h.empty: return None
    sns.set_theme(style="white", rc=_CHART_RC)
    fig, ax = plt.subplots(figsize=(16, max(8, len(h)*0.56)))
    fig.patch.set_facecolor("#FFFFFF"); ax.set_facecolor("#FFFFFF")
    try:
        cmap_obj = sns.color_palette("blend:#F7F9FC,#1B4F72", as_cmap=True)
    except Exception:
        cmap_obj = "Blues"
    mx_val = h.values.max() if h.values.size > 0 else 1
    # Build annotation with conditional coloring
    annot_text = h.map(lambda x: fmt_s(x) if x > 0 else "")
    sns.heatmap(h, ax=ax, annot=annot_text, fmt="", cmap=cmap_obj, linewidths=2.5, linecolor="#F7F9FC",
                cbar_kws={"label":"Nilai Pagu (Rp)","shrink":0.4}, annot_kws={"fontsize":10,"fontweight":"bold","color":"#0D1B2A"})
    ax.set_title(title, fontsize=18, fontweight="800", color="#0D1B2A", pad=20)
    ax.set_xlabel(""); ax.set_ylabel("")
    ax.set_yticklabels(["\n".join(textwrap.wrap(t.get_text(), 28)) for t in ax.get_yticklabels()], fontsize=11, fontweight="bold", color="#0D1B2A", rotation=0)
    ax.set_xticklabels(ax.get_xticklabels(), fontsize=12, fontweight="bold", color="#0D1B2A", rotation=20, ha="right")
    plt.tight_layout(pad=3.0)
    return fig

# ═══════════════════════════════════════════════════════════════════════════
# AGGREGATION HELPERS — DuckDB-accelerated
# ═══════════════════════════════════════════════════════════════════════════
def agg_top_pemenang(df, sektor=None, n=20, con=None):
    """Top N pemenang by total Pagu — DuckDB-accelerated."""
    if len(df) == 0:
        return pd.DataFrame()
    if con is None:
        d = df if not sektor else df[df["Sektor"] == sektor]
        if len(d) == 0: return pd.DataFrame()
        return (d.groupby("Nama_Display")
               .agg(Total_Pagu=("Pagu_Rp","sum"), Jumlah_Paket=("Pagu_Rp","count"),
                    Instansi_Unik=("Instansi_Pembeli","nunique"), Satker_Unik=("Satuan_Kerja","nunique"))
               .sort_values("Total_Pagu", ascending=False).head(n).reset_index())
    tbl = f"_tmp_pemenang_{id(df) % 99991}"
    try:
        con.execute(f"DROP VIEW IF EXISTS {tbl}")
        con.register(tbl, df)
        where = f"WHERE Sektor = '{sektor}'" if sektor else ""
        result = con.execute(f"""
            SELECT
                Nama_Display,
                SUM(Pagu_Rp)                    AS Total_Pagu,
                COUNT(*)                         AS Jumlah_Paket,
                COUNT(DISTINCT Instansi_Pembeli) AS Instansi_Unik,
                COUNT(DISTINCT Satuan_Kerja)     AS Satker_Unik
            FROM {tbl}
            {where}
            GROUP BY Nama_Display
            ORDER BY Total_Pagu DESC
            LIMIT {n}
        """).df()
        con.execute(f"DROP VIEW IF EXISTS {tbl}")
        return result
    except Exception:
        # fallback to pandas
        d = df if not sektor else df[df["Sektor"] == sektor]
        if len(d) == 0: return pd.DataFrame()
        return (d.groupby("Nama_Display")
               .agg(Total_Pagu=("Pagu_Rp","sum"), Jumlah_Paket=("Pagu_Rp","count"),
                    Instansi_Unik=("Instansi_Pembeli","nunique"), Satker_Unik=("Satuan_Kerja","nunique"))
               .sort_values("Total_Pagu", ascending=False).head(n).reset_index())


def agg_top_instansi(df, n=15, con=None):
    """Top N instansi by total Pagu — DuckDB-accelerated."""
    if len(df) == 0:
        return pd.DataFrame()
    if con is None:
        return (df.groupby("Instansi_Pembeli")
               .agg(Total_Pagu=("Pagu_Rp","sum"), Jumlah_Paket=("Pagu_Rp","count"),
                    Pemenang_Unik=("Nama_Pemenang","nunique"), Satker_Unik=("Satuan_Kerja","nunique"))
               .sort_values("Total_Pagu", ascending=False).head(n).reset_index())
    tbl = f"_tmp_instansi_{id(df) % 99991}"
    try:
        con.execute(f"DROP VIEW IF EXISTS {tbl}")
        con.register(tbl, df)
        result = con.execute(f"""
            SELECT
                Instansi_Pembeli,
                SUM(Pagu_Rp)                    AS Total_Pagu,
                COUNT(*)                         AS Jumlah_Paket,
                COUNT(DISTINCT Nama_Pemenang)    AS Pemenang_Unik,
                COUNT(DISTINCT Satuan_Kerja)     AS Satker_Unik
            FROM {tbl}
            GROUP BY Instansi_Pembeli
            ORDER BY Total_Pagu DESC
            LIMIT {n}
        """).df()
        con.execute(f"DROP VIEW IF EXISTS {tbl}")
        return result
    except Exception:
        return (df.groupby("Instansi_Pembeli")
               .agg(Total_Pagu=("Pagu_Rp","sum"), Jumlah_Paket=("Pagu_Rp","count"),
                    Pemenang_Unik=("Nama_Pemenang","nunique"), Satker_Unik=("Satuan_Kerja","nunique"))
               .sort_values("Total_Pagu", ascending=False).head(n).reset_index())


def filter_by_tema(df, tema_cfg):
    mask = pd.Series(False, index=df.index)
    for pat in tema_cfg["_re_inst"]:
        mask |= df["Instansi_Pembeli"].str.contains(pat, na=False)
    for pat in tema_cfg["_re_satker"]:
        mask |= df["Satuan_Kerja"].str.contains(pat, na=False)
    # Blacklist exclusion
    excl_mask = pd.Series(False, index=df.index)
    for pat in tema_cfg.get("_re_excl_inst", []):
        excl_mask |= df["Instansi_Pembeli"].str.contains(pat, na=False)
    return df[mask & ~excl_mask & ~df["is_pemda"]]

def filter_pemda_wilayah(df, wilayah):
    return df[df["is_pemda"] & (df["Wilayah"] == wilayah)]

def filter_by_dinas(df, dinas_name):
    pat = DINAS_COMPILED.get(dinas_name)
    if not pat: return pd.DataFrame()
    return df[df["Satuan_Kerja"].str.contains(pat, na=False)]

# ═══════════════════════════════════════════════════════════════════════════
# ★ RENDER HELPERS v4 — DuckDB-accelerated, styled download wrappers
# ═══════════════════════════════════════════════════════════════════════════
def render_top20_section(df_section, section_name, color, key_prefix, semesta=None, con=None):
    """Render Top 20 ICT + Non-ICT + Instansi + Detail — with section cards."""
    kp = re.sub(r'[^a-zA-Z0-9]', '', key_prefix)[:20]
    total_pagu = df_section["Pagu_Rp"].sum()
    n_paket = len(df_section)
    n_pemenang = df_section["Nama_Pemenang"].nunique()
    n_inst = df_section["Instansi_Pembeli"].nunique()
    n_satker = df_section["Satuan_Kerja"].nunique()
    ict_df = df_section[df_section["Sektor"]=="ICT"]
    non_df = df_section[df_section["Sektor"]=="Non-ICT"]

    # ── KPI Row ──
    c1,c2,c3,c4,c5 = st.columns(5)
    c1.markdown(kpi("Total Paket", fmt_n(n_paket)), unsafe_allow_html=True)
    c2.markdown(kpi("Total Pagu", fmt_rp(total_pagu)), unsafe_allow_html=True)
    c3.markdown(kpi("Pemenang Unik", fmt_n(n_pemenang), f"ICT: {ict_df['Nama_Pemenang'].nunique()} | Non: {non_df['Nama_Pemenang'].nunique()}"), unsafe_allow_html=True)
    c4.markdown(kpi("Instansi Pembeli", fmt_n(n_inst)), unsafe_allow_html=True)
    c5.markdown(kpi("Satuan Kerja", fmt_n(n_satker)), unsafe_allow_html=True)

    separator()

    # ── Tabs ──
    t_ict, t_non, t_inst, t_detail = st.tabs([
        f"📊 Top 20 Pemenang ICT — {section_name}",
        f"📊 Top 20 Pemenang Non-ICT — {section_name}",
        f"📊 Top 15 Instansi Pembeli — {section_name}",
        f"📋 Daftar Paket Lengkap — {section_name}"])

    with t_ict:
        agg_ict = agg_top_pemenang(df_section, "ICT", con=con)
        if len(agg_ict) > 0:
            sem_ict = ict_df["Pagu_Rp"].sum()
            fig = chart_top20(agg_ict,
                f"Top {min(20,len(agg_ict))} Pemenang ICT — {section_name}",
                f"n = {fmt_n(len(ict_df))} paket ICT  |  Total Pagu ICT: {fmt_rp(sem_ict)}  |  Semesta: {fmt_rp(semesta or total_pagu)}",
                "#1B4F72", semesta or total_pagu)
            st.pyplot(fig, use_container_width=True); plt.close(fig)

            separator()
            dl_wrap_open(f"📥 Unduh Data: Top 20 ICT — {section_name}")
            st.download_button(f"⬇️  Download Excel: Top 20 ICT — {section_name}",
                               to_excel_styled(agg_ict, f"ICT_{section_name[:15]}"),
                               f"Top20_ICT_{kp}_{datetime.now():%Y%m%d}.xlsx",
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               key=f"dlict_{kp}")
            dl_wrap_close()

            separator()
            st.markdown(f"**📋 Detail per Pemenang ICT — {section_name}** (klik untuk melihat rincian)")
            _render_pemenang_expanders(agg_ict, ict_df, f"ict{kp}")
        else:
            st.info("Tidak ada data ICT di segmen ini.")

    with t_non:
        agg_non = agg_top_pemenang(df_section, "Non-ICT", con=con)
        if len(agg_non) > 0:
            sem_non = non_df["Pagu_Rp"].sum()
            fig = chart_top20(agg_non,
                f"Top {min(20,len(agg_non))} Pemenang Non-ICT — {section_name}",
                f"n = {fmt_n(len(non_df))} paket Non-ICT  |  Total Pagu Non-ICT: {fmt_rp(sem_non)}  |  Semesta: {fmt_rp(semesta or total_pagu)}",
                "#0E6655", semesta or total_pagu)
            st.pyplot(fig, use_container_width=True); plt.close(fig)

            separator()
            dl_wrap_open(f"📥 Unduh Data: Top 20 Non-ICT — {section_name}")
            st.download_button(f"⬇️  Download Excel: Top 20 Non-ICT — {section_name}",
                               to_excel_styled(agg_non, f"NonICT_{section_name[:15]}"),
                               f"Top20_NonICT_{kp}_{datetime.now():%Y%m%d}.xlsx",
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               key=f"dlnon_{kp}")
            dl_wrap_close()

            separator()
            st.markdown(f"**📋 Detail per Pemenang Non-ICT — {section_name}** (klik untuk melihat rincian)")
            _render_pemenang_expanders(agg_non, non_df, f"non{kp}")
        else:
            st.info("Tidak ada data Non-ICT di segmen ini.")

    with t_inst:
        agg_inst = agg_top_instansi(df_section, con=con)
        if len(agg_inst) > 0:
            agg_inst_chart = agg_inst.rename(columns={"Instansi_Pembeli":"Nama_Display","Pemenang_Unik":"Instansi_Unik"})
            fig = chart_top20(agg_inst_chart,
                f"Top {min(15,len(agg_inst))} Instansi Pembeli — {section_name}",
                f"{fmt_n(n_inst)} instansi total  |  Total Pagu: {fmt_rp(total_pagu)}",
                "#B7770D", total_pagu, figsize=(15, 7))
            st.pyplot(fig, use_container_width=True); plt.close(fig)
        else:
            st.info("Tidak ada data instansi.")

    with t_detail:
        cols_show = ["Nama_Pemenang","Prediksi_Nama","Pagu_Rp","Instansi_Pembeli","Satuan_Kerja","Lokasi","Metode_Pemilihan","Sektor","Nama_Paket"]
        cols_avail = [c for c in cols_show if c in df_section.columns]
        df_show = df_section[cols_avail].sort_values("Pagu_Rp", ascending=False).head(500)
        df_disp = df_show.copy()
        if "Pagu_Rp" in df_disp.columns: df_disp["Pagu_Rp"] = df_disp["Pagu_Rp"].apply(fmt_rp)

        st.markdown(f"**TABEL: Daftar {fmt_n(min(500,len(df_section)))} Paket Terbesar — {section_name}**")
        st.dataframe(df_disp, use_container_width=True, hide_index=True, height=400)

        separator()
        dl_wrap_open(f"📥 Unduh Data: Detail Paket — {section_name}")
        st.download_button(f"⬇️  Download Excel: Detail Paket — {section_name}",
                           to_excel_styled(df_show.head(2000), f"Detail_{section_name[:15]}"),
                           f"Detail_{kp}_{datetime.now():%Y%m%d}.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key=f"dldet_{kp}")
        dl_wrap_close()


def _render_pemenang_expanders(agg_df, detail_df, key_prefix):
    for idx, (_, row) in enumerate(agg_df.head(20).iterrows()):
        si = row["Nama_Display"]
        dsi = detail_df[detail_df["Nama_Display"] == si]
        n_inst = int(row.get("Instansi_Unik", 0))
        n_satk = int(row.get("Satker_Unik", 0))
        val = row["Total_Pagu"]
        header = f"**#{idx+1} {si}** — {fmt_rp(val)} | {n_inst} instansi • {n_satk} satker • {int(row['Jumlah_Paket'])} paket"
        with st.expander(header, expanded=False):
            if len(dsi) == 0: st.caption("Detail tidak tersedia."); continue
            k1, k2, k3, k4 = st.columns(4)
            k1.markdown(kpi("Total Pagu", fmt_rp(dsi["Pagu_Rp"].sum())), unsafe_allow_html=True)
            k2.markdown(kpi("Jumlah Paket", fmt_n(len(dsi))), unsafe_allow_html=True)
            k3.markdown(kpi("Instansi Unik", fmt_n(dsi["Instansi_Pembeli"].nunique())), unsafe_allow_html=True)
            k4.markdown(kpi("Satker Unik", fmt_n(dsi["Satuan_Kerja"].nunique())), unsafe_allow_html=True)
            col_a, col_b = st.columns(2)
            with col_a:
                st.markdown(f"**Instansi Pembeli — {si[:30]}**")
                ia = (dsi.groupby("Instansi_Pembeli").agg(Pagu=("Pagu_Rp","sum"), Paket=("Pagu_Rp","count"))
                      .sort_values("Pagu", ascending=False).reset_index())
                ia["Pagu"] = ia["Pagu"].apply(fmt_rp); ia.columns = ["Instansi", "Total Pagu", "Paket"]
                st.dataframe(ia, use_container_width=True, hide_index=True, height=200)
            with col_b:
                st.markdown(f"**Satuan Kerja (Top 15) — {si[:30]}**")
                sa = (dsi.groupby("Satuan_Kerja").agg(Pagu=("Pagu_Rp","sum"), Paket=("Pagu_Rp","count"))
                      .sort_values("Pagu", ascending=False).head(15).reset_index())
                sa["Pagu"] = sa["Pagu"].apply(fmt_rp); sa.columns = ["Satuan Kerja", "Total Pagu", "Paket"]
                st.dataframe(sa, use_container_width=True, hide_index=True, height=200)
            st.markdown(f"**Daftar Paket — {si[:30]}**")
            cs = [c for c in ["Nama_Paket","Pagu_Rp","Instansi_Pembeli","Satuan_Kerja","Lokasi"] if c in dsi.columns]
            dp = dsi[cs].sort_values("Pagu_Rp", ascending=False).head(100).copy()
            dp["Pagu_Rp"] = dp["Pagu_Rp"].apply(fmt_rp)
            st.dataframe(dp, use_container_width=True, hide_index=True, height=220)


def render_drilldown_table(df_sub, label, key_prefix, con=None):
    kp = re.sub(r'[^a-zA-Z0-9]', '', key_prefix)[:25]
    cols_show = ["Nama_Pemenang","Prediksi_Nama","Pagu_Rp","Instansi_Pembeli","Satuan_Kerja","Lokasi","Metode_Pemilihan","Sektor","Nama_Paket"]
    cols_avail = [c for c in cols_show if c in df_sub.columns]
    df_raw = df_sub[cols_avail].sort_values("Pagu_Rp", ascending=False)

    ci, cn = st.columns(2)
    with ci:
        st.markdown(f"**Top 20 Pemenang ICT — {label[:30]}**")
        ai = agg_top_pemenang(df_sub, "ICT", 20, con=con)
        if len(ai) > 0:
            ai_d = ai.copy(); ai_d["Total_Pagu"] = ai_d["Total_Pagu"].apply(fmt_rp)
            st.dataframe(ai_d, use_container_width=True, hide_index=True)
        else: st.caption("Tidak ada pemenang ICT.")
    with cn:
        st.markdown(f"**Top 20 Pemenang Non-ICT — {label[:30]}**")
        an = agg_top_pemenang(df_sub, "Non-ICT", 20, con=con)
        if len(an) > 0:
            an_d = an.copy(); an_d["Total_Pagu"] = an_d["Total_Pagu"].apply(fmt_rp)
            st.dataframe(an_d, use_container_width=True, hide_index=True)
        else: st.caption("Tidak ada pemenang Non-ICT.")

    separator()
    st.markdown(f"**Daftar Paket Lengkap — {label}** ({fmt_n(len(df_raw))} paket)")
    df_disp = df_raw.head(1000).copy()
    df_disp["Pagu_Rp"] = df_disp["Pagu_Rp"].apply(fmt_rp)
    st.dataframe(df_disp, use_container_width=True, hide_index=True, height=350)

    separator()
    dl_wrap_open(f"📥 Unduh Data — {label}")
    c1, c2 = st.columns(2)
    with c1:
        st.download_button(f"⬇️  Download CSV — {label[:25]}",
                           df_raw.head(5000).to_csv(index=False).encode("utf-8"),
                           f"{kp}_{datetime.now():%Y%m%d}.csv", "text/csv", key=f"dlcsv_{kp}")
    with c2:
        st.download_button(f"⬇️  Download Excel — {label[:25]}",
                           to_excel_styled(df_raw.head(5000), label[:25]),
                           f"{kp}_{datetime.now():%Y%m%d}.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=f"dlxls_{kp}")
    dl_wrap_close()


# ═══════════════════════════════════════════════════════════════════════════
# MAIN — LOAD DATA
# ═══════════════════════════════════════════════════════════════════════════
st.markdown("""
<div class="hero">
    <h1>📊 Dashboard Realisasi Pengadaan Pemerintah — INAPROC 2025</h1>
    <p>Data Realisasi Seluruh KLPD &nbsp;|&nbsp; Telkomsel Enterprise — Bid Management Intelligence</p>
</div>""", unsafe_allow_html=True)

df, total_rows, n_matched, err, db_source = load_and_process()
if err:
    st.error(f"⚠️ {err}")
    st.stop()

# ═══════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ═══════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("<div style='text-align:center;padding:20px 0;border-bottom:1px solid rgba(255,255,255,0.08)'>"
                "<span style='font-size:24px'>📊</span><br>"
                "<span style='font-family:DM Sans,sans-serif;font-size:16px;font-weight:800;color:#FFF!important'>INAPROC INTELLIGENCE</span><br>"
                f"<span style='font-family:Inter,sans-serif;font-size:10px;color:#5D8AA8!important'>Sumber: {db_source}</span></div>",
                unsafe_allow_html=True)
    st.markdown("---")
    st.markdown("<div style='background:rgba(14,102,85,0.20);color:#52D9B9!important;border-radius:8px;"
                f"padding:8px 12px;font-family:Inter,sans-serif;font-size:12px;font-weight:600;text-align:center'>"
                f"✅ {fmt_n(len(df))} paket aktif</div>", unsafe_allow_html=True)
    pq = _get_parquet_path()
    if os.path.exists(pq):
        st.caption(f"⚡ Cache aktif ({os.path.getsize(pq)/1e6:.0f} MB)")
    st.markdown("---")
    st.markdown("<p style='font-size:9px!important;font-weight:700!important;color:#5D8AA8!important;"
                "letter-spacing:2px!important;text-transform:uppercase!important;margin-bottom:4px!important'>NAVIGASI</p>",
                unsafe_allow_html=True)
    view = st.radio("📌 MENU",
                    ["🏛️ Analisis per Kementerian/Lembaga",
                     "🗺️ Analisis per Wilayah & Dinas",
                     "🔍 Pencarian & Filter Data"], index=0, label_visibility="collapsed")
    st.markdown("---")
    sektor_filter = st.radio("🔍 Filter Sektor", ["Semua","ICT","Non-ICT"], index=0)
    st.markdown("---")
    if os.path.exists(pq):
        if st.button("🗑️ Rebuild Cache", use_container_width=True):
            os.remove(pq); st.cache_data.clear(); st.rerun()
    if st.button("🚪 Logout", use_container_width=True):
        st.session_state["authenticated"] = False; st.rerun()
    st.caption(f"Telkomsel Enterprise\nEBPM — Data Science\n{datetime.now():%d %B %Y}")

dff = df[df["Sektor"] == sektor_filter] if sektor_filter != "Semua" else df
duck_con = get_duckdb_conn()

# ═══════════════════════════════════════════════════════════════════════════
# VIEW: PER K/L (6 TEMA STRATEGIS)
# ═══════════════════════════════════════════════════════════════════════════
if "Kementerian" in view:
    section_open("🏛️ ANALISIS PER KEMENTERIAN / LEMBAGA — 6 TEMA STRATEGIS",
                 "Pemenang pengadaan per tema K/L (exclude Pemda Kab/Kota/Provinsi)", "blue")

    tabs = st.tabs(list(TEMA_KL.keys()))
    for tab, tema_name in zip(tabs, TEMA_KL.keys()):
        with tab:
            cfg = TEMA_KL[tema_name]; df_tema = filter_by_tema(dff, cfg)
            st.markdown(f"""
            <div class="rcard" style="background:#FAFAFA;border-color:{cfg['color']}">
                <h3 style="color:{cfg['color']}!important">{cfg['icon']} {tema_name}</h3>
                <p>{cfg['desc']} &nbsp;|&nbsp; <strong>{fmt_n(len(df_tema))} paket</strong> &nbsp;|&nbsp; <strong>{fmt_rp(df_tema['Pagu_Rp'].sum())}</strong></p>
            </div>""", unsafe_allow_html=True)
            if len(df_tema) == 0: st.warning("Tidak ada data matching untuk tema ini."); continue
            tema_kp = re.sub(r'[^a-zA-Z0-9]', '', tema_name)[:10]
            render_top20_section(df_tema, tema_name, cfg["color"], f"kl{tema_kp}",
                                semesta=dff["Pagu_Rp"].sum(), con=duck_con)
            separator_thick()
            st.markdown(f"**🔎 Drill-down: Pilih Instansi Pembeli — {tema_name}**")
            inst_list = df_tema["Instansi_Pembeli"].value_counts().head(20).index.tolist()
            if inst_list:
                sel_inst = st.selectbox(f"Pilih Instansi ({tema_name})", inst_list, key=f"selinst_{tema_kp}")
                if sel_inst:
                    df_inst = df_tema[df_tema["Instansi_Pembeli"] == sel_inst]
                    st.markdown(f"**{sel_inst}** — {fmt_n(len(df_inst))} paket | {fmt_rp(df_inst['Pagu_Rp'].sum())}")
                    inst_kp = re.sub(r'[^a-zA-Z0-9]', '', sel_inst)[:10]
                    render_drilldown_table(df_inst, sel_inst, f"kldrill{tema_kp}{inst_kp}", con=duck_con)


# ═══════════════════════════════════════════════════════════════════════════
# VIEW: PER WILAYAH & DINAS
# ═══════════════════════════════════════════════════════════════════════════
elif "Wilayah" in view:
    section_open("🗺️ ANALISIS PER WILAYAH & DINAS STRATEGIS",
                 "Pemenang pengadaan dari Pemda (Kab/Kota/Provinsi) per wilayah dan dinas", "red")

    sub_view = st.radio("Pilih Analisis:", [
        "📍 Per Wilayah (6 Wilayah Indonesia)",
        "📡 Diskominfo se-Indonesia (Khusus ICT)",
        "🏢 Per Dinas Strategis per Wilayah"],
        horizontal=True, key="wil_subview")

    if "Per Wilayah" in sub_view:
        pemda_by_wil = {w: filter_pemda_wilayah(dff, w) for w in WILAYAH_LIST}
        wdata = [w for w in WILAYAH_LIST if len(pemda_by_wil[w]) > 0]
        sem_all = sum(pemda_by_wil[w]["Pagu_Rp"].sum() for w in WILAYAH_LIST)
        if not wdata: st.warning("Tidak ada data Pemda.")
        else:
            tabs = st.tabs([f"{W_CFG[w]['i']} Wilayah {w}" for w in wdata])
            for tab, w in zip(tabs, wdata):
                with tab:
                    cf = W_CFG[w]; dw = pemda_by_wil[w]
                    w_kp = re.sub(r'[^a-zA-Z0-9]', '', w)[:8]
                    st.markdown(f"""
                    <div class="rcard" style="background:{cf['bg']};border-color:{cf['c']}">
                        <h3 style="color:{cf['c']}!important">{cf['i']} Wilayah {w} — Pengadaan Pemda</h3>
                        <p>{fmt_n(len(dw))} paket &nbsp;|&nbsp; {fmt_rp(dw['Pagu_Rp'].sum())} &nbsp;|&nbsp;
                        {fmt_n(dw['Nama_Pemenang'].nunique())} pemenang &nbsp;|&nbsp; {fmt_n(dw['Instansi_Pembeli'].nunique())} instansi Pemda</p>
                    </div>""", unsafe_allow_html=True)
                    render_top20_section(dw, f"Wilayah {w}", cf["c"], f"wil{w_kp}", semesta=sem_all, con=duck_con)
                    separator_thick()
                    st.markdown(f"**🔎 Drill-down: Pilih Instansi Pemda — Wilayah {w}**")
                    inst_list = dw["Instansi_Pembeli"].value_counts().head(25).index.tolist()
                    if inst_list:
                        sel = st.selectbox("Pilih Instansi Pemda", inst_list, key=f"selwil{w_kp}")
                        if sel:
                            df_sel = dw[dw["Instansi_Pembeli"] == sel]
                            st.markdown(f"**{sel}** — {fmt_n(len(df_sel))} paket | {fmt_rp(df_sel['Pagu_Rp'].sum())}")
                            render_drilldown_table(df_sel, sel, f"wildrill{w_kp}{re.sub(r'[^a-zA-Z0-9]','',sel)[:10]}", con=duck_con)

    elif "Diskominfo" in sub_view:
        section_open("📡 ANALISIS DISKOMINFO SE-INDONESIA — KHUSUS PEMDA",
                     "Satuan kerja Diskominfo / Komunikasi & Informatika dari seluruh Pemda Indonesia", "blue")
        df_dkom = filter_by_dinas(dff, "Diskominfo")
        df_dkom = df_dkom[df_dkom["is_pemda"]]
        if len(df_dkom) == 0: st.warning("Tidak ada data DISKOMINFO Pemda.")
        else:
            c1,c2,c3,c4 = st.columns(4)
            c1.markdown(kpi("Total Paket", fmt_n(len(df_dkom))), unsafe_allow_html=True)
            c2.markdown(kpi("Total Pagu", fmt_rp(df_dkom["Pagu_Rp"].sum())), unsafe_allow_html=True)
            c3.markdown(kpi("Instansi Pemda", fmt_n(df_dkom["Instansi_Pembeli"].nunique())), unsafe_allow_html=True)
            c4.markdown(kpi("Pemenang Unik", fmt_n(df_dkom["Nama_Pemenang"].nunique())), unsafe_allow_html=True)
            render_top20_section(df_dkom, "Diskominfo se-Indonesia", "#4A235A", "dkom",
                                semesta=dff["Pagu_Rp"].sum(), con=duck_con)
            separator_thick()
            st.markdown("**🔎 Drill-down: Pilih Instansi Pemda (Diskominfo)**")
            dkom_inst = df_dkom["Instansi_Pembeli"].value_counts().head(30)
            sel_dkom = st.selectbox("Pilih Instansi Pemda", dkom_inst.index.tolist(), key="seldkominst")
            if sel_dkom:
                df_sel = df_dkom[df_dkom["Instansi_Pembeli"] == sel_dkom]
                st.markdown(f"**{sel_dkom}** — {fmt_n(len(df_sel))} paket | {fmt_rp(df_sel['Pagu_Rp'].sum())}")
                render_drilldown_table(df_sel, sel_dkom, f"dkomdrill{re.sub(r'[^a-zA-Z0-9]','',sel_dkom)[:12]}", con=duck_con)

    elif "Per Dinas" in sub_view:
        section_open("🏢 ANALISIS PER DINAS STRATEGIS — PEMDA",
                     "Pilih wilayah → pilih dinas → lihat pemenang per instansi Pemda", "plain")
        sel_w = st.selectbox("Pilih Wilayah:", WILAYAH_LIST, key="seldinaswil")
        cf = W_CFG[sel_w]; ws = WILAYAH_STRATEGY[sel_w]
        dw = filter_pemda_wilayah(dff, sel_w)
        w_kp = re.sub(r'[^a-zA-Z0-9]', '', sel_w)[:8]
        if len(dw) == 0: st.warning(f"Tidak ada data Pemda di {sel_w}.")
        else:
            dinas_tabs = st.tabs([f"🏢 {d} — {sel_w}" for d in ws["dinas"]])
            for d_idx, (dtab, dinas_name) in enumerate(zip(dinas_tabs, ws["dinas"])):
                with dtab:
                    df_dinas = filter_by_dinas(dw, dinas_name)
                    d_kp = re.sub(r'[^a-zA-Z0-9]', '', dinas_name)[:8]
                    unique_kp = f"ds{d_idx}{d_kp}{w_kp}"
                    if len(df_dinas) == 0: st.info(f"Tidak ada data {dinas_name} di {sel_w}."); continue
                    st.markdown(f"""
                    <div class="rcard" style="background:{cf['bg']};border-color:{cf['c']}">
                        <h3 style="color:{cf['c']}!important">{dinas_name} — Wilayah {sel_w}</h3>
                        <p>{fmt_n(len(df_dinas))} paket &nbsp;|&nbsp; {fmt_rp(df_dinas['Pagu_Rp'].sum())} &nbsp;|&nbsp;
                        {fmt_n(df_dinas['Nama_Pemenang'].nunique())} pemenang &nbsp;|&nbsp; {fmt_n(df_dinas['Instansi_Pembeli'].nunique())} instansi</p>
                    </div>""", unsafe_allow_html=True)
                    render_top20_section(df_dinas, f"{dinas_name} — {sel_w}", cf["c"], unique_kp,
                                        semesta=dw["Pagu_Rp"].sum(), con=duck_con)
                    separator_thick()
                    st.markdown(f"**🔎 Drill-down: Pilih Instansi ({dinas_name} — {sel_w})**")
                    inst_list = df_dinas["Instansi_Pembeli"].value_counts().head(20).index.tolist()
                    if inst_list:
                        sel = st.selectbox("Pilih Instansi", inst_list, key=f"sel{unique_kp}")
                        if sel:
                            df_si = df_dinas[df_dinas["Instansi_Pembeli"] == sel]
                            st.markdown(f"**{sel}** — {fmt_n(len(df_si))} paket | {fmt_rp(df_si['Pagu_Rp'].sum())}")
                            render_drilldown_table(df_si, sel, f"dsdrill{unique_kp}{re.sub(r'[^a-zA-Z0-9]','',sel)[:10]}", con=duck_con)


# ═══════════════════════════════════════════════════════════════════════════
# VIEW: PENCARIAN & FILTER DATA
# ═══════════════════════════════════════════════════════════════════════════
elif "Pencarian" in view:
    section_open("🔍 PENCARIAN & FILTER DATA PENGADAAN",
                 "Cari berdasarkan nama pemenang, instansi, wilayah, atau sektor", "plain")

    c1, c2, c3, c4 = st.columns(4)
    with c1: search_pem = st.text_input("🔍 Cari Nama Pemenang", key="search_pem", placeholder="Ketik nama vendor...")
    with c2: search_inst = st.text_input("🏛️ Cari Instansi", key="search_inst", placeholder="Ketik instansi...")
    with c3: sel_wil = st.selectbox("🗺️ Pilih Wilayah", ["Semua"] + WILAYAH_LIST, key="exp_wil")
    with c4: sel_sek = st.selectbox("🔍 Pilih Sektor", ["Semua","ICT","Non-ICT"], key="exp_sek")

    df_exp = dff.copy()
    if search_pem: df_exp = df_exp[df_exp["Nama_Pemenang"].str.contains(search_pem, case=False, na=False)]
    if search_inst: df_exp = df_exp[df_exp["Instansi_Pembeli"].str.contains(search_inst, case=False, na=False)]
    if sel_wil != "Semua": df_exp = df_exp[df_exp["Wilayah"] == sel_wil]
    if sel_sek != "Semua": df_exp = df_exp[df_exp["Sektor"] == sel_sek]

    st.markdown(f"**Hasil Pencarian:** {fmt_n(len(df_exp))} paket | {fmt_rp(df_exp['Pagu_Rp'].sum())} | "
                f"{fmt_n(df_exp['Nama_Pemenang'].nunique())} pemenang unik")

    if len(df_exp) > 0:
        separator()
        section_open(f"📊 Top 20 Pemenang — Hasil Pencarian ({fmt_n(len(df_exp))} paket)",
                     f"Total Pagu: {fmt_rp(df_exp['Pagu_Rp'].sum())}", "plain")

        agg_exp = agg_top_pemenang(df_exp, n=20, con=duck_con)
        if len(agg_exp) > 0:
            fig = chart_top20(agg_exp,
                f"Top 20 Pemenang — Hasil Pencarian",
                f"n = {fmt_n(len(df_exp))} paket  |  Total: {fmt_rp(df_exp['Pagu_Rp'].sum())}",
                "#C8102E", df_exp["Pagu_Rp"].sum())
            st.pyplot(fig, use_container_width=True); plt.close(fig)

        separator()
        section_open(f"📋 Data Mentah — Hasil Pencarian (maks. 1.000 baris)", "", "plain")

        cols_show = ["Nama_Pemenang","Pagu_Rp","Instansi_Pembeli","Satuan_Kerja","Lokasi","Wilayah","Sektor","Metode_Pemilihan","Jenis_Pengadaan","Nama_Paket"]
        cols_avail = [c for c in cols_show if c in df_exp.columns]
        df_disp = df_exp[cols_avail].sort_values("Pagu_Rp", ascending=False).head(1000).copy()
        df_disp_fmt = df_disp.copy()
        df_disp_fmt["Pagu_Rp"] = df_disp_fmt["Pagu_Rp"].apply(fmt_rp)
        st.dataframe(df_disp_fmt, use_container_width=True, hide_index=True, height=500)

        separator()
        dl_wrap_open("📥 Unduh Data: Hasil Pencarian (maks. 5.000 baris)")
        st.download_button("⬇️  Download Excel: Hasil Pencarian",
                           to_excel_styled(df_exp[cols_avail].sort_values("Pagu_Rp", ascending=False).head(5000), "Pencarian"),
                           f"Pencarian_{datetime.now():%Y%m%d}.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_explorer")
        dl_wrap_close()
    else:
        st.info("Tidak ada data yang cocok dengan filter pencarian.")


# ═══════════════════════════════════════════════════════════════════════════
# FOOTER
# ═══════════════════════════════════════════════════════════════════════════
st.markdown(f"""
<div class="footer-box">
    <p style="font-family:'DM Sans',sans-serif;font-size:15px;font-weight:700;margin:0 0 8px">
        Dashboard Realisasi Pengadaan Pemerintah — INAPROC 2025</p>
    <p class="ft-sub" style="font-family:'Inter',sans-serif;font-size:13px;margin:0 0 6px">
        Telkomsel Enterprise &nbsp;|&nbsp; EBPM — EPES — Direktorat PnT &nbsp;|&nbsp; Bid Management Data Science</p>
    <p class="ft-dim" style="font-family:'Inter',sans-serif;font-size:11px;margin:0">
        📊 {fmt_n(len(df))} paket dari {fmt_n(total_rows)} records &nbsp;|&nbsp; Sumber: {db_source}
        &nbsp;|&nbsp; Generated: {datetime.now():%d %B %Y %H:%M}</p>
</div>""", unsafe_allow_html=True)